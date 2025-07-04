from datetime import datetime
import calendar
import openpyxl

MONTHS = [
    '',
    'Jan',
    'Feb',
    'Mar',
    'Apr',
    'May',
    'Jun',
    'Jul',
    'Aug',
    'Sep',
    'Oct',
    'Nov',
    'Dec',        
]

replacements = {
    "MONTH START": None,
    "MONTH END": None,
    
    'YEAR': None,

    'MONTH': None,
    'MONTH-NUM': None,

    'MONTH-1': None,
    'MONTH-1-NUM': None,

    'MONTH-2': None,
    'MONTH-2-NUM': None,

    'YEAR-2': None
}

def generate_xlsx(ltg: int):
    month = datetime.now().month
    year = datetime.now().year

    days_in_month = calendar.monthrange(year, month)[1]

    # one month ago
    month_1 = 12 if month == 1 else month - 1

    # two months ago
    month_2 = (10 + month) if month <= 2 else month - 2


    with open(f'template.txt', 'r') as f:
        template = f.read()
    template = template.replace('LTG *', f'LTG {ltg}')

    # Populate automated fields
    replacements['MONTH START'] = f"{month}/1/{year}"    
    replacements['MONTH END'] = f"{month}/{days_in_month}/{year}"   
    replacements['MONTH'] = MONTHS[month]
    replacements['MONTH-NUM'] = month
    replacements['YEAR'] = year

    replacements['MONTH-1'] = MONTHS[month_1]
    replacements['MONTH-1-NUM'] = month_1

    replacements['MONTH-2'] = MONTHS[month_2]
    replacements['MONTH-2-NUM'] = month_2

    replacements['YEAR-1'] =  year - 1 if month <= 1 else year
    replacements['YEAR-2'] =  year - 1 if month <= 2 else year

    wb = openpyxl.load_workbook(f"input.xlsx", data_only=True)
    ws = wb.active

    for i in range(1, 100):
        title = ws.cell(i, 1).value
        value = ws.cell(i, 2).value
        default_value = ws.cell(i, 3).value

        if value is None or value == '' :
            value = default_value
        
        if isinstance(value, float):
            value = round(value, 2)
            value = f"${value}"
        elif isinstance(value, str) and value.strip().endswith('00:00:00'):
            value = value.replace('00:00:00', '').strip()
            value_lst = value.split('-')
            value = '/'.join([value_lst[1], value_lst[2], value_lst[0]])
        elif isinstance(value, datetime):
            value = value.strftime("%m/%d/%Y")
        else:
            print(value, type(value))

        if value is not None:
            template = template.replace(f"[[{title}]]", str(value))

    for k, v in replacements.items():
        template = template.replace(f"[[{k}]]", str(v))
    


    with open(f'output-ltg-{ltg}.txt', 'w') as f:
        f.write(template)

def main():
    for i in range(1, 3):
        generate_xlsx(i)

if __name__ == '__main__':
    main()